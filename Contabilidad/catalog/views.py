from django.http.response import HttpResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.core.files.storage import FileSystemStorage
import pandas as pd
import os
from Contabilidad.settings import MEDIA_ROOT
import zipfile
from openpyxl import load_workbook
import mysql.connector
from .mysql_connection import mysqlconection
import shutil



@login_required
def home(request):
    return render(request,'registration/inicio.html')

def process_excel_file(excel_path, existing_facturas):
    df = pd.read_excel(excel_path)
    wb = load_workbook(excel_path)
    sheet = wb.active

    rows_to_delete = [row[0].row for row in sheet.iter_rows() if any(isinstance(cell.value, str) and "Application response" in cell.value for cell in row)]

    for row_idx in reversed(rows_to_delete):
        sheet.delete_rows(row_idx)

    wb.save(excel_path)
    df = pd.read_excel(excel_path)

    found_data_list = []
    not_found_data_list = []

    for index, row in df.iterrows():
        if pd.isnull(row['Prefijo']):
            row['Prefijo'] = ''

        concatenated_value = f"{row['Prefijo']}{row['Folio']}"

        if concatenated_value in existing_facturas:
            print(f"Factura {concatenated_value} encontrada en la base de datos.")
            found_data = {'NIT Emisor': row['NIT Emisor'],
                          'Factura': concatenated_value,
                          'Nombre Emisor': row['Nombre Emisor'],
                          'Fecha Recepción': row['Fecha Recepción'],
                          'CUFE/CUDE': row['CUFE/CUDE']}
            found_data_list.append(found_data)
        else:
            print(f"Factura {concatenated_value} no encontrada en la base de datos.")
            not_found_data = {'NIT Emisor': row['NIT Emisor'],
                              'Factura': concatenated_value,
                              'Nombre Emisor': row['Nombre Emisor'],
                              'Fecha Recepción': row['Fecha Recepción'],
                              'CUFE/CUDE': row['CUFE/CUDE']}
            not_found_data_list.append(not_found_data)

    found_df = pd.DataFrame(found_data_list)
    not_found_df = pd.DataFrame(not_found_data_list)

    found_file_path = os.path.join('catalog','media', 'convertido', 'encontradas.xlsx')
    not_found_file_path = os.path.join('catalog','static', 'resultados', 'no_encontradas.xlsx')

    # Guardar ambos archivos
    found_df.to_excel(found_file_path, index=False, engine='openpyxl')
    not_found_df.to_excel(not_found_file_path, index=False, engine='openpyxl')
    
    return df

def Upload_zip(request):
    if request.method == 'POST' and 'file' in request.FILES:
        uploaded_file = request.FILES['file']
        fs = FileSystemStorage()
        filename = fs.save(uploaded_file.name, uploaded_file)
        zip_file_path = os.path.join(MEDIA_ROOT, filename)
        extract_folder = os.path.join(MEDIA_ROOT, 'extracted')

        with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
            zip_ref.extractall(extract_folder)

        excel_files = [f for f in os.listdir(extract_folder) if f.endswith('.xlsx')]
        if excel_files:
            excel_path = os.path.join(extract_folder, excel_files[0])

            mysql_connection = mysqlconection()

            if mysql_connection:
                try:
                    if mysql_connection.is_connected():
                        print("Conexión MySQL establecida")

                        cursor = mysql_connection.cursor()
                        cursor.execute("SELECT nombre_factura FROM facturas")
                        result_set = cursor.fetchall()
                        existing_facturas = [row[0] for row in result_set]
                        cursor.close()

                        df = process_excel_file(excel_path, existing_facturas)

                except mysql.connector.Error as mysql_error:
                    print("Error de MySQL:", mysql_error)
                finally:
                    if mysql_connection and mysql_connection.is_connected():
                        mysql_connection.close()
                        print("Conexión MySQL cerrada")

            else:
                print('Falló MySQL_Connection')

            return render(request, 'registration/Upload_zip.html', {'filename': excel_files[0], 'data': df.to_html()})
        else:
            os.rmdir(extract_folder)
            return HttpResponse("No se encontró un archivo Excel (.xlsx) dentro del archivo zip.")

    return render(request, 'registration/Upload_zip.html')

def reiniciarSistema(request):
    # Eliminar y recrear el directorio /media/convertido/
    convertido_dir = "catalog/media/"
    media_dir = "catalog/media/convertido"
    
    shutil.rmtree(convertido_dir, ignore_errors=True)
    os.makedirs(media_dir)

    return render(request, "registration/sistemaReiniciado.html")

